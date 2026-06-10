import json
import time
import os
import random
from datetime import datetime
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from curl_cffi import requests 

class HiringCafeScraper:
    def __init__(self, cf_clearance_cookie=None):
        self.base_url = "https://hiring.cafe"
        self.session = requests.Session()
        
        # Aligned exactly to Chrome 148 / Windows 11 footprint parameters
        USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
        
        self.headers = {
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Accept-Language": "en,en-US;q=0.9",
            "Sec-Ch-Ua": '"Not/A)Brand";v="8", "Chromium";v="148", "Google Chrome";v="148"',
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Upgrade-Insecure-Requests": "1",
            "Referer": "https://hiring.cafe/",
            "Origin": "https://hiring.cafe"
        }
        
        if cf_clearance_cookie:
            self.session.cookies.set("cf_clearance", cf_clearance_cookie)
            
        self.build_id = None
        self.search_state = None
        self.current_tech = "General"

    def bootstrap(self):
        """Extracts live Build ID strings securely using active cookie profiles."""
        print("[*] Syncing context properties with website home portal...")
        
        res = self.session.get(self.base_url, headers=self.headers, impersonate="chrome", timeout=30)
        
        if res.status_code != 200:
            print(f"[!] Core handshake failed. Server status flag: {res.status_code}")
            if "challenge" in res.text.lower() or "turnstile" in res.text.lower():
                print("[!] Notification: Local security clearance verification expired or was missing.")
            raise Exception("Cloudflare blocking routine intercepted standard stream processing.")

        soup = BeautifulSoup(res.text, 'html.parser')
        script_tag = soup.find('script', id='__NEXT_DATA__')
        
        if not script_tag:
            raise Exception("Parsing anomaly: __NEXT_DATA__ block structure missing from raw response.")
            
        data = json.loads(script_tag.string)
        self.build_id = data.get('buildId')
        self.search_state = data['props']['pageProps'].get('initialSearchState', {})
        print(f"[+] Operational channel initialized. Active Build ID string: {self.build_id}")

    def apply_filters(self, query=None, remote_only=False, min_salary=None, days_back=1, country_code=None, languages=None, simple_apps_only=False):
        """Builds custom filter states matching targeted payload specifications."""
        if not self.search_state: self.bootstrap()
        
        if query:
            self.current_tech = query
            self.search_state['searchQuery'] = query

        self.search_state['dateFetchedPastNDays'] = days_back
        self.search_state['sortBy'] = "date" 
        
        if remote_only:
            self.search_state['workplaceTypes'] = ["Remote"]
        else:
            self.search_state['workplaceTypes'] = ["Remote", "Hybrid", "Onsite"]

        if languages:
            self.search_state['languageRequirements'] = [lang.lower() for lang in languages]

        # Application Complexity Matrix filtering argument
        if simple_apps_only:
            self.search_state['applicationEase'] = ["Simple Application Forms"]
        else:
            if 'applicationEase' in self.search_state:
                del self.search_state['applicationEase']

        if min_salary:
            self.search_state.update({
                "minCompensationLowEnd": min_salary,
                "minCompensationHighEnd": min_salary,
                "currency": {"label": "USD", "value": "USD"},
                "frequency": {"label": "Yearly", "value": "Yearly"},
                "calcFrequency": "Yearly",
                "restrictJobsToTransparentSalaries": True
            })

        if country_code == "US":
            self.search_state['locations'] = [{
                "id": "FxY1yZQBoEtHp_8UEq7V",
                "types": ["country"],
                "formatted_address": "United States",
                "address_components": [{"long_name": "United States", "short_name": "US", "types": ["country"]}]
            }]
        elif country_code == "PK":
             self.search_state['locations'] = [{"formatted_address": "Pakistan"}]

    def clean_html(self, raw_html):
        if not raw_html: return ""
        return BeautifulSoup(raw_html, "html.parser").get_text(separator=' ').strip()

    def get_jobs(self, pages=1, skip_ats_list=None):
        """Fetches dynamic JSON listings and handles custom programmatic blacklists."""
        if not self.build_id: self.bootstrap()
        all_data = []
        scrape_date = datetime.now().strftime("%Y-%m-%d")

        api_headers = self.headers.copy()
        api_headers.update({
            "Accept": "application/json, text/plain, */*",
            "x-nextjs-data": "1"
        })

        for p in range(pages):
            api_url = f"{self.base_url}/_next/data/{self.build_id}/index.json"
            params = {"searchState": json.dumps(self.search_state), "page": p}
            
            print(f"[*] Extracting page indices {p} for query reference '{self.current_tech}'...")
            res = self.session.get(api_url, params=params, headers=api_headers, impersonate="chrome")
            
            if res.status_code != 200: 
                print(f"[!] Execution halted at pagination index {p}. Server state: {res.status_code}")
                break
                

            data = res.json()
            hits = data['pageProps'].get('ssrHits', [])
            if not hits: 
                print("[!] Zero matching target records found on current page layer.")
                break

           
            

            for hit in hits:
                ats_source = str(hit.get('source', '')).lower()
                apply_url = str(hit.get('apply_url', '')).lower()
                
                # Exclude platform elements matching target blacklists
                if skip_ats_list:
                    if any(ats.lower() in ats_source or ats.lower() in apply_url for ats in skip_ats_list):
                        continue

                info = hit.get('job_information', {})
                processed = hit.get('v5_processed_job_data', {})
                
                # --- UPDATED PARSING MATRIX FROM YOUR LOGS ---
                raw_description = (
                    processed.get('requirements_summary') or 
                    hit.get('description_text') or 
                    hit.get('job_description_text') or
                    info.get('description') or 
                    hit.get('description') or ""
                )

                all_data.append({
                    "Date": scrape_date,
                    "Platform": "Hiring Cafe",
                    "Site": "Hiring Cafe",
                    "Apply Link": hit.get('apply_url'),
                    "Company Name": processed.get('company_name') or hit.get('company_name'),
                    "Job Title": info.get('title') or hit.get('title'),
                    "Company Link": processed.get('company_website'),
                    "Job Description": self.clean_html(raw_description),
                    "Tech Stack": self.current_tech
                })

            wait = random.uniform(4.5, 7.5)
            print(f"[*] Completed stream page {p}. Cooling processor thread for {wait:.1f}s...")
            time.sleep(wait)
            
        return all_data

def save_to_excel_with_gap(new_data, file_path):
    if not new_data: return
    new_df = pd.DataFrame(new_data)
    sheet_name = 'Jobs'
    cols_order = ["Number", "Date", "Platform", "Site", "Apply Link", "Company Name", "Job Title", "Company Link", "Job Description", "Tech Stack"]
    
    if os.path.exists(file_path):
        existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
        last_num = existing_df['Number'].dropna().max() if 'Number' in existing_df.columns else 0
        new_df.insert(0, 'Number', range(int(last_num) + 1, int(last_num) + 1 + len(new_df)))
        new_df = new_df.reindex(columns=cols_order, fill_value="")
        
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = load_workbook(file_path)
            start_row = writer.book[sheet_name].max_row + 2
            new_df.to_excel(writer, startrow=start_row, index=False, header=False, sheet_name=sheet_name)
    else:
        new_df.insert(0, 'Number', range(1, 1 + len(new_df)))
        new_df = new_df.reindex(columns=cols_order, fill_value="")
        new_df.to_excel(file_path, index=False, sheet_name=sheet_name)

# ==========================================
# RUNTIME ENGINE INITIALIZATION BLOCK
# ==========================================
if __name__ == "__main__":
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    OUTPUT_FILE = os.path.join(BASE_DIR, "Master_Job_Tracker.xlsx")

    # 1. Open Chrome -> Go to hiring.cafe -> F12 DevTools -> Application -> Cookies
    # 2. Extract and paste your valid 'cf_clearance' value below:
    MY_CF_COOKIE = "7Jw2ulmiSY01HmyaTba71HK8ThpXWm04CXzGqZ1IefA-1779297131-1.2.1.1-VAQEoMOqv7AB0mOBQpZGmw5cXU5Mf5PLXstuY.nOOzRJTOPUqExGH89V_H2HgJ9eJM_9VBbsZzZCkV6XO3DU5Gx3r4UOXRGGrKrhIM.JtkIBxYNHE1phY1ifkHBj._fiFxa_m7l48rMeTTQGLRbqrLDlKSpYfLsIdi24CKoARRMujDWOsZ5O9uDbp2tB.8E622eRUzXCyrBWdXutfCvFm_JfQXVzDQJMxZfCtjxZJsD7Qt7od0G2i62XncPaLJ4CXlBnG_769Z8g2a7Vj.NZF8Yg06AqkZERNfbVS3_zo5m.tbF_BFyd.3_KK57CPp8_84sDKt7k10velSD5iNirBYlj.xo2SqySdOyo.KDH1DJWVtV4QdBKe.wzeoYrXvcU1IGgLgfNUsOUtdl62HtNq2lOmHdpmux5JpRFDtpccgg" 

    if MY_CF_COOKIE == "PASTE_YOUR_LIVE_CHROME_148_COOKIE_HERE" or not MY_CF_COOKIE:
        print("[!] Warning: Global execution token variable 'MY_CF_COOKIE' is empty.")

    scraper = HiringCafeScraper(cf_clearance_cookie=MY_CF_COOKIE)

    # Scraper Task Iteration Queue Matrix
    tasks = [
        {"tech": "Machine learning engineer", "lang": ["English"], "remote": True, "country": "US"}
    ]

    for task in tasks:
        try:
            scraper.apply_filters(
                query=task['tech'], 
                languages=task['lang'], 
                remote_only=task['remote'], 
                country_code=task['country'],
                days_back=1,
                simple_apps_only=True # Isolates Greenhouse/Ashby/Lever types natively
            )
            
            # Explicitly strips out lengthy application platform engines on output conversion  "workday", "taleo", "icims"
            results = scraper.get_jobs(pages=1, skip_ats_list=[])
            
            save_to_excel_with_gap(results, OUTPUT_FILE)
            print(f"[✔] Batch indexing successfully added for stack context: {task['tech']}")
            
        except Exception as e:
            print(f"[✘] Session compilation aborted: {e}")

    print(f"\n[COMPLETE] Operations closed. Matrix saved at: {OUTPUT_FILE}")